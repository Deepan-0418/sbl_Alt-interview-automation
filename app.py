from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify
from datetime import datetime
from db import init_db, insert_user, insert_typing_result
from pdf_utils import generate_typing_test_pdf, generate_error_report_pdf
from werkzeug.utils import escape
import random
from time import time
import json
import os
import logging
import sys
import zipfile
import io
import openpyxl
from openpyxl.styles import Font, Alignment
from werkzeug.utils import secure_filename

# ── Logging ────────────────────────────────────────────────────
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# ── Path setup: PyInstaller bundle vs normal Python / Render ───
if getattr(sys, 'frozen', False):
    # Running as a PyInstaller .exe
    import shutil
    base_path = sys._MEIPASS
    user_dir  = os.path.join(os.path.expanduser("~"), "InterviewAutomation")
    if not os.path.exists(user_dir):
        os.makedirs(user_dir)

    writable_data_json = os.path.join(user_dir, 'data.json')
    writable_static    = os.path.join(user_dir, 'static')

    if not os.path.exists(writable_data_json):
        shutil.copy(os.path.join(base_path, 'data.json'), writable_data_json)
    if not os.path.exists(writable_static):
        shutil.copytree(os.path.join(base_path, 'static'), writable_static)

else:
    # Normal Python — local dev or Render deployment
    base_path = os.path.dirname(__file__)

    # DATA_ROOT: set to /data on Render (persistent disk), else use project root
    data_root = os.environ.get('DATA_ROOT', base_path)
    user_dir  = data_root

    writable_data_json = os.path.join(data_root, 'data.json')
    writable_static    = os.path.join(base_path, 'static')  # always use repo static

    # On first Render boot, /data/data.json won't exist — copy from repo
    if not os.path.exists(writable_data_json):
        import shutil
        src = os.path.join(base_path, 'data.json')
        if os.path.exists(src):
            os.makedirs(os.path.dirname(writable_data_json), exist_ok=True)
            shutil.copy(src, writable_data_json)

# ── Flask app ──────────────────────────────────────────────────
app = Flask(__name__, static_folder=writable_static, static_url_path='/static')
app.config['SECRET_KEY']         = os.environ.get('SECRET_KEY', 'fallback-dev-key-change-this')
app.config['UPLOAD_FOLDER']      = writable_static
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'xlsx'}

# ── Load data.json ─────────────────────────────────────────────
data_file_path = writable_data_json

SAMPLE_PARAGRAPHS     = {'easy': [], 'medium': [], 'hard': []}
HANDWRITTEN_TEXTS     = []
EXCEL_QUIZ_QUESTIONS  = []
EXCEL_PRACTICAL_TASKS = []

if not os.path.exists(data_file_path):
    logger.error(f"data.json not found at: {data_file_path}")
else:
    try:
        with open(data_file_path, 'r', encoding='utf-8') as f:
            data = json.loads(f.read())
        SAMPLE_PARAGRAPHS     = data.get('sample_paragraphs',    {'easy': [], 'medium': [], 'hard': []})
        HANDWRITTEN_TEXTS     = data.get('handwritten_texts',    [])
        EXCEL_QUIZ_QUESTIONS  = data.get('excel_quiz_questions', [])
        EXCEL_PRACTICAL_TASKS = data.get('excel_practical_tasks',[])
    except json.JSONDecodeError as e:
        logger.error(f"JSON parsing error in data.json: {e}")
    except Exception as e:
        logger.error(f"Unexpected error loading data.json: {e}")

# ── Jinja filter: cache-busting timestamp ─────────────────────
app.jinja_env.filters['timestamp'] = lambda _: str(int(time()))

# ── Admin credentials ──────────────────────────────────────────
# Username: set via ADMIN_USERNAME env var, defaults to 'admin'
# Password: checked fresh at every login — defaults to today's date
#           in YYYYMMDD format (e.g. 20260318), auto-updates daily.
#           Override by setting ADMIN_PASSWORD env var on Render.
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')

def get_admin_password():
    """
    Returns the admin password checked fresh on every login.
    Default: today's date as YYYYMMDD — updates automatically at midnight.
    Override: set the ADMIN_PASSWORD environment variable for a fixed password.
    """
    return os.environ.get('ADMIN_PASSWORD', 'Admin@' + datetime.now().strftime('%Y%m%d'))

# ── Initialize database ────────────────────────────────────────
with app.app_context():
    init_db()

# ── Typing attempt config ──────────────────────────────────────
ATTEMPT_CONFIG = [
    {'label': 'Warm-Up',        'difficulty': 'easy',   'time_limit': 300, 'scored': False},
    {'label': 'First Attempt',  'difficulty': 'easy',   'time_limit': 120, 'scored': True},
    {'label': 'Second Attempt', 'difficulty': 'medium', 'time_limit': 120, 'scored': True},
    {'label': 'Third Attempt',  'difficulty': 'hard',   'time_limit': 120, 'scored': True},
]


# ══════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


def generate_excel_template():
    """Generate an Excel template with task descriptions from EXCEL_PRACTICAL_TASKS."""
    workbook = openpyxl.Workbook()

    sheet1  = workbook.active; sheet1.title = "Function"
    sheet2  = workbook.create_sheet("Sort")
    sheet3  = workbook.create_sheet("Replace")
    sheet4  = workbook.create_sheet("Concatenate")
    sheet5  = workbook.create_sheet("Sum & Average")
    sheet6  = workbook.create_sheet("Insert Row & Delete Column")
    sheet7  = workbook.create_sheet("Trim & Length")
    sheet8  = workbook.create_sheet("Left & Right")
    sheet9  = workbook.create_sheet("Count")
    sheet10 = workbook.create_sheet("Duplicates")

    all_sheets = [sheet1, sheet2, sheet3, sheet4, sheet5,
                  sheet6, sheet7, sheet8, sheet9, sheet10]

    for sheet in all_sheets:
        sheet['A1'] = "Excel Practical Test Instructions"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.merge_cells('A1:C1')

    if EXCEL_PRACTICAL_TASKS:
        random.seed(42)
        tasks = EXCEL_PRACTICAL_TASKS[:]
        random.shuffle(tasks)
        total_sheets = len(all_sheets)
        for idx, task in enumerate(tasks):
            sheet = all_sheets[idx % total_sheets]
            row   = (idx // total_sheets) + 2
            sheet[f'A{row}'] = f"Task {task['task_id']}: {task['description']}"
            sheet[f'A{row}'].alignment = Alignment(wrap_text=True)

        sheet1['B2'] = "Sample Data"
        sheet1['B2'].font = Font(bold=True)
        for row in range(3, 12):
            sheet1[f'B{row}'] = random.randint(1, 100)

    for sheet in all_sheets:
        sheet.column_dimensions['A'].width = 60
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15

    template_path = os.path.join(app.config['UPLOAD_FOLDER'], 'excel_practical_template.xlsx')
    workbook.save(template_path)
    return template_path


def validate_excel_against_master(user_file_path, master_file_path):
    """Compare user's Excel file against master; 1 per sheet if identical, else 0."""
    sheet_names = [
        "Function", "Sort", "Replace", "Concatenate", "Sum & Average",
        "Insert Row & Delete Column", "Trim & Length", "Left & Right",
        "Count", "Duplicates"
    ]
    try:
        user_wb   = openpyxl.load_workbook(user_file_path,   data_only=True)
        master_wb = openpyxl.load_workbook(master_file_path, data_only=True)

        if (not all(n in user_wb.sheetnames   for n in sheet_names) or
                not all(n in master_wb.sheetnames for n in sheet_names)):
            logger.error("Sheet mismatch between user and master files")
            return 0.0, {n: 0 for n in sheet_names}

        sheet_scores = {}
        for sheet_name in sheet_names:
            u = user_wb[sheet_name]
            m = master_wb[sheet_name]

            if u.max_row != m.max_row or u.max_column != m.max_column:
                sheet_scores[sheet_name] = 0
                continue

            def norm(v):
                if isinstance(v, (int, float)):
                    return str(round(float(v), 2))
                return str(v).strip() if v is not None else ""

            identical = all(
                norm(u.cell(r, c).value) == norm(m.cell(r, c).value)
                for r in range(1, u.max_row + 1)
                for c in range(1, u.max_column + 1)
            )
            sheet_scores[sheet_name] = 1 if identical else 0
            logger.info(f"Sheet {sheet_name}: {'Identical' if identical else 'Different'}")

        overall = round((sum(sheet_scores.values()) / len(sheet_names)) * 100, 2)
        logger.info(f"Overall Quality Score: {overall}%")
        return overall, sheet_scores

    except Exception as e:
        logger.error(f"Error validating Excel file: {e}")
        return 0.0, {n: 0 for n in sheet_names}


# ══════════════════════════════════════════════════════════════
# ADMIN ROUTES
# ══════════════════════════════════════════════════════════════

@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        # Password checked fresh on every login — updates automatically each day
        if username == ADMIN_USERNAME and password == get_admin_password():
            session['admin_logged_in'] = True
            flash('Login successful!', 'success')
            return redirect(url_for('admin_dashboard'))

        flash('Invalid credentials.', 'error')
        return redirect(url_for('admin_login'))

    return render_template('admin_login.html')


@app.route('/admin_dashboard', methods=['GET', 'POST'])
def admin_dashboard():
    if not session.get('admin_logged_in'):
        flash('Please log in as admin.', 'error')
        return redirect(url_for('admin_login'))

    global SAMPLE_PARAGRAPHS, HANDWRITTEN_TEXTS, EXCEL_QUIZ_QUESTIONS, EXCEL_PRACTICAL_TASKS

    if request.method == 'POST':

        # ── Update data.json ───────────────────────────────────
        if 'data_json' in request.form:
            try:
                new_data = json.loads(request.form.get('data_json'))
                if not new_data.get('excel_practical_tasks'):
                    flash('Warning: excel_practical_tasks is empty or missing.', 'warning')
                with open(data_file_path, 'w', encoding='utf-8') as f:
                    json.dump(new_data, f, indent=4)
                SAMPLE_PARAGRAPHS     = new_data.get('sample_paragraphs',    {'easy': [], 'medium': [], 'hard': []})
                HANDWRITTEN_TEXTS     = new_data.get('handwritten_texts',    [])
                EXCEL_QUIZ_QUESTIONS  = new_data.get('excel_quiz_questions', [])
                EXCEL_PRACTICAL_TASKS = new_data.get('excel_practical_tasks',[])
                generate_excel_template()
                flash('data.json updated successfully!', 'success')
            except json.JSONDecodeError:
                flash('Invalid JSON format.', 'error')
            except Exception as e:
                flash(f'Error updating data.json: {str(e)}', 'error')

        # ── Upload handwritten image ───────────────────────────
        elif 'file' in request.files:
            file             = request.files['file']
            handwritten_text = request.form.get('handwritten_text')
            if file and allowed_file(file.filename) and handwritten_text:
                filename  = secure_filename(file.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)
                with open(data_file_path, 'r', encoding='utf-8') as f:
                    d = json.load(f)
                d['handwritten_texts'].append({'image': filename, 'text': handwritten_text})
                with open(data_file_path, 'w', encoding='utf-8') as f:
                    json.dump(d, f, indent=4)
                HANDWRITTEN_TEXTS.append({'image': filename, 'text': handwritten_text})
                flash('Image uploaded and data.json updated!', 'success')
            else:
                flash('Invalid file or text input.', 'error')

        # ── Delete image ───────────────────────────────────────
        elif 'delete_image' in request.form:
            image_to_delete = request.form.get('delete_image')
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], image_to_delete)
            if os.path.exists(file_path):
                os.remove(file_path)
                with open(data_file_path, 'r', encoding='utf-8') as f:
                    d = json.load(f)
                d['handwritten_texts'] = [
                    i for i in d['handwritten_texts']
                    if i['image'] != image_to_delete
                ]
                with open(data_file_path, 'w', encoding='utf-8') as f:
                    json.dump(d, f, indent=4)
                HANDWRITTEN_TEXTS = d['handwritten_texts']
                flash('Image deleted successfully!', 'success')
            else:
                flash('Image not found.', 'error')

        return redirect(url_for('admin_dashboard'))

    with open(data_file_path, 'r', encoding='utf-8') as f:
        data_json = json.load(f)

    image_files = [
        fn for fn in os.listdir(app.config['UPLOAD_FOLDER'])
        if allowed_file(fn)
    ]
    return render_template(
        'admin_dashboard.html',
        data_json=json.dumps(data_json, indent=4),
        image_files=image_files,
    )


@app.route('/admin_logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    flash('Logged out successfully.', 'success')
    return redirect(url_for('admin_login'))


@app.route('/admin_clear_session')
def admin_clear_session():
    if not session.get('admin_logged_in'):
        flash('Please log in as admin.', 'error')
        return redirect(url_for('admin_login'))
    keys_to_remove = [k for k in list(session.keys()) if k != 'admin_logged_in']
    for key in keys_to_remove:
        session.pop(key)
    flash('Candidate session cleared successfully.', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/debug_static_files')
def debug_static_files():
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Admin access required'}), 403
    return jsonify({'static_files': os.listdir(app.config['UPLOAD_FOLDER'])})


@app.route('/debug_tasks')
def debug_tasks():
    return jsonify({'excel_practical_tasks': EXCEL_PRACTICAL_TASKS})


# ══════════════════════════════════════════════════════════════
# CANDIDATE ROUTES
# ══════════════════════════════════════════════════════════════

@app.route('/')
def welcome():
    return render_template('welcome.html')


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        name           = escape(request.form.get('name',           '').strip())
        location       = escape(request.form.get('location',       '').strip())
        distance       = request.form.get('distance',       '').strip()
        attempt_number = request.form.get('attempt_number', '').strip()
        dob            = request.form.get('dob',            '').strip()
        signup_date    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if not all([name, location, distance, attempt_number, dob]):
            flash('Please fill out all fields.', 'error')
            return redirect(url_for('signup'))

        try:
            float(distance)
        except ValueError:
            flash('Distance must be a valid number.', 'error')
            return redirect(url_for('signup'))

        if attempt_number not in ['1st', '2nd', '3rd']:
            flash('Attempt Number must be 1st, 2nd, or 3rd.', 'error')
            return redirect(url_for('signup'))

        try:
            dob_date = datetime.strptime(dob, '%Y-%m-%d')
            today    = datetime.now()
            age      = today.year - dob_date.year - (
                (today.month, today.day) < (dob_date.month, dob_date.day)
            )
            if age < 18:
                flash('You must be at least 18 years old to sign up.', 'error')
                return redirect(url_for('signup'))
        except ValueError:
            flash('Invalid date of birth format.', 'error')
            return redirect(url_for('signup'))

        selected_excel_questions = (
            random.sample(EXCEL_QUIZ_QUESTIONS, min(10, len(EXCEL_QUIZ_QUESTIONS)))
            if EXCEL_QUIZ_QUESTIONS else []
        )

        session.update({
            'user_name':                 name,
            'location':                  location,
            'distance':                  distance,
            'attempt_number':            attempt_number,
            'dob':                       dob,
            'signup_date':               signup_date,
            'handwritten_completed':     False,
            'typing_completed':          False,
            'typing_attempts':           0,
            'typing_results':            [],
            'excel_quiz_completed':      False,
            'excel_practical_completed': False,
            'current_image_index':       0,
            'handwritten_results':       [],
            'excel_quiz_results':        [],
            'excel_quiz_questions':      [],
            'selected_excel_questions':  selected_excel_questions,
            'excel_practical_results':   [],
            'excel_practical_file':      None,
            'excel_practical_score':     None,
            'excel_sheet_scores':        None,
        })

        insert_user(name, signup_date, location, distance, attempt_number, dob)
        flash('Sign-up successful! Please complete all four rounds.', 'success')
        return redirect(url_for('welcome'))

    return render_template('signup.html')


@app.route('/handwritten_round', methods=['GET'])
def handwritten_round():
    name        = session.get('user_name')
    signup_date = session.get('signup_date')
    if not name or not signup_date:
        flash('Please sign up first.', 'error')
        return redirect(url_for('signup'))

    if session.get('handwritten_completed'):
        if (session.get('typing_completed') and
                session.get('excel_quiz_completed') and
                session.get('excel_practical_completed')):
            return redirect(url_for('thank_you'))
        return render_template('handwritten_round.html',
                               name=name, signup_date=signup_date, completed=True)

    if not HANDWRITTEN_TEXTS:
        flash('No handwritten texts available.', 'error')
        return redirect(url_for('welcome'))

    if 'selected_handwritten_texts' not in session:
        session['selected_handwritten_texts'] = random.sample(
            HANDWRITTEN_TEXTS, min(10, len(HANDWRITTEN_TEXTS))
        )
        session['current_image_index'] = 0

    current_index              = session.get('current_image_index', 0)
    selected_handwritten_texts = session.get('selected_handwritten_texts', [])

    if not selected_handwritten_texts or current_index >= len(selected_handwritten_texts):
        flash('No handwritten texts available.', 'error')
        return redirect(url_for('welcome'))

    current_image    = selected_handwritten_texts[current_index]['image']
    session.modified = True

    return render_template('handwritten_round.html',
                           name=name, signup_date=signup_date,
                           image=current_image, completed=False)


@app.route('/submit_handwritten', methods=['POST'])
def submit_handwritten():
    name = session.get('user_name')
    if not name:
        return jsonify({'error': 'Please sign up first.'}), 403

    user_input                 = escape(request.form.get('handwritten_input', '').strip())
    selected_handwritten_texts = session.get('selected_handwritten_texts', [])
    if not selected_handwritten_texts:
        return jsonify({'error': 'No handwritten texts available.'}), 403

    index = session.get('current_image_index', 0)
    if index >= len(selected_handwritten_texts):
        return jsonify({'error': 'No handwritten texts available.'}), 403

    current    = selected_handwritten_texts[index]
    is_correct = user_input == current['text']
    status     = 'Correct' if is_correct else 'Incorrect'

    session['handwritten_results'] = session.get('handwritten_results', [])
    session['handwritten_results'].append({
        'image':        current['image'],
        'status':       status,
        'user_input':   user_input,
        'correct_text': current['text'],
    })

    next_index = index + 1
    if next_index < len(selected_handwritten_texts):
        session['current_image_index'] = next_index
        next_image     = selected_handwritten_texts[next_index]['image']
        next_image_url = url_for('static', filename=next_image, _external=True)
    else:
        session['handwritten_completed'] = True
        flash('All handwritten verifications completed!', 'success')
        next_image     = current['image']
        next_image_url = url_for('static', filename=next_image, _external=True)

    session.modified = True

    response = {
        'completed':      session.get('handwritten_completed', False),
        'next_image':     next_image,
        'next_image_url': next_image_url,
    }
    if response['completed']:
        response['redirect'] = (
            url_for('thank_you')
            if (session.get('typing_completed') and
                session.get('excel_quiz_completed') and
                session.get('excel_practical_completed'))
            else url_for('welcome')
        )
    return jsonify(response)


@app.route('/typing_test', methods=['GET', 'POST'])
def typing_test():
    name        = session.get('user_name')
    signup_date = session.get('signup_date')
    if not name or not signup_date:
        flash('Please sign up first.', 'error')
        return redirect(url_for('signup'))

    if 'typing_attempts' not in session:
        session['typing_attempts'] = 0
        session['typing_results']  = []

    attempt_index = session.get('typing_attempts', 0)

    if session.get('typing_completed', False):
        flash('Typing test completed.', 'success')
        if (session.get('handwritten_completed') and
                session.get('excel_quiz_completed') and
                session.get('excel_practical_completed')):
            return redirect(url_for('thank_you'))
        return redirect(url_for('welcome'))

    if request.method == 'POST':
        config    = ATTEMPT_CONFIG[attempt_index]
        is_scored = config['scored']

        user_name          = escape(request.form.get('user_name',          '').strip())
        selected_paragraph = escape(request.form.get('selected_paragraph', '').strip())
        difficulty         = config['difficulty']
        time_limit         = config['time_limit']

        try:
            wpm      = float(request.form.get('wpm',      0))
            accuracy = float(request.form.get('accuracy', 0))
        except (ValueError, TypeError):
            flash('Invalid test data.', 'error')
            return redirect(url_for('typing_test'))

        test_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if is_scored:
            insert_typing_result(user_name, wpm, accuracy, time_limit)
            session['typing_results'].append({
                'wpm':                wpm,
                'accuracy':           accuracy,
                'time_limit':         time_limit,
                'test_date':          test_date,
                'selected_paragraph': selected_paragraph,
                'difficulty':         difficulty,
                'attempt':            attempt_index,
            })

        session['typing_attempts'] = attempt_index + 1
        next_index = attempt_index + 1

        if next_index >= len(ATTEMPT_CONFIG):
            session['typing_completed'] = True
            flash('Typing test completed! All attempts done.', 'success')
            session.modified = True
            if (session.get('handwritten_completed') and
                    session.get('excel_quiz_completed') and
                    session.get('excel_practical_completed')):
                return redirect(url_for('thank_you'))
            return redirect(url_for('welcome'))

        next_config = ATTEMPT_CONFIG[next_index]
        if is_scored:
            flash(f'{config["label"]} submitted! Now starting {next_config["label"]}.', 'success')
        else:
            flash(f'Warm-up done! Now starting {next_config["label"]} — this is scored.', 'info')

        session.modified = True
        return redirect(url_for('typing_test'))

    # GET
    attempt_index = session.get('typing_attempts', 0)
    if attempt_index >= len(ATTEMPT_CONFIG):
        session['typing_completed'] = True
        return redirect(url_for('welcome'))

    config     = ATTEMPT_CONFIG[attempt_index]
    difficulty = config['difficulty']
    time_limit = config['time_limit']
    paragraphs = SAMPLE_PARAGRAPHS.get(difficulty, [])

    if not paragraphs:
        flash(f'No paragraphs available for {difficulty} difficulty.', 'error')
        return redirect(url_for('welcome'))

    selected_paragraph = random.choice(paragraphs)

    return render_template(
        'typing_test.html',
        name=name,
        signup_date=signup_date,
        paragraph=selected_paragraph,
        difficulty=difficulty,
        time_limit=time_limit,
        attempt_label=config['label'],
        is_warmup=not config['scored'],
        attempt_index=attempt_index,
        typing_attempts=len(session.get('typing_results', [])),
        typing_results=session.get('typing_results', []),
    )


@app.route('/excel_quiz', methods=['GET', 'POST'])
def excel_quiz():
    name        = session.get('user_name')
    signup_date = session.get('signup_date')
    if not name or not signup_date:
        flash('Please sign up first.', 'error')
        return redirect(url_for('signup'))

    if session.get('excel_quiz_completed'):
        if (session.get('typing_completed') and
                session.get('handwritten_completed') and
                session.get('excel_practical_completed')):
            return redirect(url_for('thank_you'))
        return render_template('excel_quiz.html',
                               name=name, signup_date=signup_date, completed=True)

    if request.method == 'POST':
        answers   = request.form.to_dict()
        questions = session.get('excel_quiz_questions', [])
        if not questions:
            flash('No quiz questions available.', 'error')
            return redirect(url_for('welcome'))

        results = []
        score   = 0
        for i, question in enumerate(questions):
            user_answer = answers.get(f'question_{i}', '')
            is_correct  = user_answer == question['correct']
            if is_correct:
                score += 1
            results.append({
                'question':       question['question'],
                'user_answer':    user_answer,
                'correct_answer': question['correct'],
                'status':         'Correct' if is_correct else 'Incorrect',
            })

        session['excel_quiz_results']   = results
        session['excel_quiz_completed'] = True
        session['excel_quiz_score']     = score
        session['excel_quiz_total']     = len(questions)
        session.modified = True

        flash(f'Excel quiz completed! You scored {score}/{len(questions)}.', 'success')
        if (session.get('handwritten_completed') and
                session.get('typing_completed') and
                session.get('excel_practical_completed')):
            return redirect(url_for('thank_you'))
        return redirect(url_for('welcome'))

    selected_questions = session.get('selected_excel_questions', [])
    if not selected_questions:
        flash('No quiz questions available.', 'error')
        return redirect(url_for('welcome'))

    shuffled = selected_questions.copy()
    random.shuffle(shuffled)
    session['excel_quiz_questions'] = shuffled
    session.modified = True

    return render_template(
        'excel_quiz.html',
        name=name,
        signup_date=signup_date,
        questions=list(enumerate(shuffled)),
        completed=False,
    )


@app.route('/excel_practical', methods=['GET', 'POST'])
def excel_practical():
    name        = session.get('user_name')
    signup_date = session.get('signup_date')
    if not name or not signup_date:
        flash('Please sign up first.', 'error')
        return redirect(url_for('signup'))

    if not EXCEL_PRACTICAL_TASKS:
        flash('No tasks available for Excel practical test.', 'error')
        return redirect(url_for('welcome'))

    if session.get('excel_practical_completed'):
        if (session.get('typing_completed') and
                session.get('handwritten_completed') and
                session.get('excel_quiz_completed')):
            return redirect(url_for('thank_you'))
        return render_template(
            'excel_practical.html',
            name=name, signup_date=signup_date, completed=True,
            tasks=EXCEL_PRACTICAL_TASKS,
            quality_score=session.get('excel_practical_score'),
            sheet_scores=session.get('excel_sheet_scores'),
        )

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded.', 'error')
            return redirect(url_for('excel_practical'))

        file = request.files['file']
        if file.filename == '' or not allowed_file(file.filename):
            flash('Invalid file format. Please upload an .xlsx file.', 'error')
            return redirect(url_for('excel_practical'))

        sanitized_name = "".join(
            c for c in name if c.isalnum() or c == '_'
        ).strip()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename  = f"excel_practical_{sanitized_name}_{timestamp}.xlsx"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        try:
            master_file_path = os.path.join(
                app.config['UPLOAD_FOLDER'], 'master_excel_solution.xlsx'
            )
            if not os.path.exists(master_file_path):
                flash('Master Excel file not found. Please contact the administrator.', 'error')
                os.remove(file_path)
                return redirect(url_for('excel_practical'))

            quality_score, sheet_scores = validate_excel_against_master(
                file_path, master_file_path
            )

            session['excel_practical_file']      = filename
            session['excel_practical_score']     = quality_score
            session['excel_sheet_scores']        = sheet_scores
            session['excel_practical_completed'] = True
            session.modified = True

            flash(f'File uploaded! Overall Quality Score: {quality_score}%', 'success')
            if (session.get('handwritten_completed') and
                    session.get('typing_completed') and
                    session.get('excel_quiz_completed')):
                return redirect(url_for('thank_you'))
            return redirect(url_for('excel_practical'))

        except Exception as e:
            logger.error(f"Error processing Excel file: {e}")
            flash(f'Error processing uploaded file: {str(e)}', 'error')
            if os.path.exists(file_path):
                os.remove(file_path)
            return redirect(url_for('excel_practical'))

    return render_template(
        'excel_practical.html',
        name=name, signup_date=signup_date,
        completed=False, tasks=EXCEL_PRACTICAL_TASKS,
        quality_score=None, sheet_scores=None,
    )


@app.route('/download_excel_template')
def download_excel_template():
    template_path = os.path.join(
        app.config['UPLOAD_FOLDER'], 'excel_practical_template.xlsx'
    )
    if not os.path.exists(template_path):
        flash('Excel template file not found. Please contact the administrator.', 'error')
        return redirect(url_for('excel_practical'))
    try:
        candidate_name = session.get('user_name', '')
        dob            = session.get('dob', '')
        attempt        = session.get('attempt_number', '')

        safe_name = (
            "".join(c for c in candidate_name if c.isalnum() or c == ' ')
            .strip().replace(' ', '_')
        )
        try:
            dob_fmt = datetime.strptime(dob, '%Y-%m-%d').strftime('%d%m%Y')
        except ValueError:
            dob_fmt = ''

        parts = ['excel_practical_template']
        if safe_name: parts.append(safe_name)
        if dob_fmt:   parts.append(dob_fmt)
        if attempt:   parts.append(attempt.replace(' ', ''))

        download_name = '_'.join(parts) + '.xlsx'
        return send_file(template_path, as_attachment=True, download_name=download_name)

    except Exception as e:
        logger.error(f"Error serving Excel template: {e}")
        flash('Error downloading the template file. Please try again.', 'error')
        return redirect(url_for('excel_practical'))


@app.route('/thank_you')
def thank_you():
    name        = session.get('user_name')
    signup_date = session.get('signup_date')
    if not name or not signup_date:
        flash('Please sign up first.', 'error')
        return redirect(url_for('signup'))

    if not (session.get('handwritten_completed') and
            session.get('typing_completed') and
            session.get('excel_quiz_completed') and
            session.get('excel_practical_completed')):
        flash('Please complete all four tests.', 'error')
        return redirect(url_for('welcome'))

    return render_template(
        'thank_you.html',
        name=name,
        signup_date=signup_date,
        results=session.get('typing_results', []),
        handwritten_results=session.get('handwritten_results', []),
        excel_quiz_results=session.get('excel_quiz_results', []),
        excel_score=session.get('excel_quiz_score', 0),
        excel_total=session.get('excel_quiz_total', 0),
        excel_practical_file=session.get('excel_practical_file'),
        excel_practical_tasks=EXCEL_PRACTICAL_TASKS,
        excel_practical_score=session.get('excel_practical_score'),
        excel_sheet_scores=session.get('excel_sheet_scores', {}),
    )


@app.route('/download_results')
def download_results():
    name        = session.get('user_name')
    signup_date = session.get('signup_date')
    if not (name and
            session.get('typing_completed') and
            session.get('handwritten_completed') and
            session.get('excel_quiz_completed') and
            session.get('excel_practical_completed')):
        flash('No results available. Please complete all tests.', 'error')
        return redirect(url_for('welcome'))

    dob            = session.get('dob', '')
    location       = session.get('location', '')
    distance       = session.get('distance', 0.0)
    attempt_number = session.get('attempt_number', '')

    try:
        sanitized_date = datetime.strptime(
            signup_date, '%Y-%m-%d %H:%M:%S'
        ).strftime('%Y%m%d_%H%M%S')
    except ValueError:
        sanitized_date = 'unknown_date'

    results_buffer, results_filename = generate_typing_test_pdf(
        name=name,
        typing_results=session.get('typing_results', []),
        handwritten_results=session.get('handwritten_results', []),
        excel_quiz_results=session.get('excel_quiz_results', []),
        excel_score=session.get('excel_quiz_score', 0),
        excel_total=session.get('excel_quiz_total', 0),
        excel_practical_file=session.get('excel_practical_file'),
        excel_practical_tasks=EXCEL_PRACTICAL_TASKS,
        excel_practical_score=session.get('excel_practical_score'),
        excel_sheet_scores=session.get('excel_sheet_scores'),
        location=location,
        distance=distance,
        attempt_number=attempt_number,
        signup_date=signup_date,
        dob=dob,
    )

    error_buffer, error_filename = generate_error_report_pdf(
        name=name,
        handwritten_results=session.get('handwritten_results', []),
        excel_quiz_results=session.get('excel_quiz_results', []),
        signup_date=signup_date,
        dob=dob,
    )

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(results_filename, results_buffer.getvalue())
        zf.writestr(error_filename,   error_buffer.getvalue())
        excel_practical_file = session.get('excel_practical_file')
        if excel_practical_file:
            fp = os.path.join(app.config['UPLOAD_FOLDER'], excel_practical_file)
            if os.path.exists(fp):
                zf.write(fp, excel_practical_file)

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name=f"Results_{name}_{sanitized_date}.zip",
        mimetype='application/zip',
    )


@app.route('/clear_session')
def clear_session():
    session.clear()
    flash('Session cleared. Please sign up again.', 'info')
    return redirect(url_for('signup'))


# ══════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════

if __name__ == '__main__':
    app.run(
        host='0.0.0.0',
        port=int(os.environ.get('PORT', 5000)),
        debug=os.environ.get('FLASK_DEBUG', 'false').lower() == 'true',
    )